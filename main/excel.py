import datetime
import json
import shutil

import xlrd
import xlwt
from django.db.models import Q

import openpyxl
from django.db import models
from django.utils.dateparse import parse_date
from loguru import logger
from xlwt import Worksheet

from Egida import settings
from accounts.models import SchoolUser
from main.models import *
from main.models.PersonalModel import Director
from main.models.services import get_model_fields

START_ROW = 4
START_COLUMN = 1


def create_new_schools_and_users_from_excel(file_path):
    dir = file_path
    rb = xlrd.open_workbook(dir)
    sheet = rb.sheet_by_index(0)
    for i in range(4, sheet.nrows):
        done = int(float(i) / float(sheet.nrows) * 100)
        logger.info(str.format("IMPORT DONE ON {0}%", done))
        values = sheet.row_values(i)
        INN = int(values[0])
        district = values[1]
        name = values[2]
        shortname = values[3]
        try:
            phone = int(values[4])
        except:
            phone = values[4]
        address = values[5]
        try:
            district = District.objects.get(name=district)
        except:
            district = District.objects.create(name=district)
        try:
            school = School.objects.create(INN=INN, name=name, shortname=shortname, phone=phone, address=address,
                                           district=district)
            SchoolUser.objects.create(username=INN, password=INN, school=school)
        except BaseException as ex:
            logger.debug(ex)


def update_schools_from_excel():
    dir = settings.BASE_DIR.__str__() + "\Spisok_OU_06_10_20.xlsx"
    rb = xlrd.open_workbook(dir)
    sheet = rb.sheet_by_index(0)
    for i in range(4, sheet.nrows):
        logger.info(str.format("import done {0}", int(float(i) / float(sheet.nrows) * 100)))
        values = sheet.row_values(i)
        INN = int(values[0])
        district = values[1]
        name = values[2]
        shortname = values[3]
        try:
            phone = int(values[4])
        except:
            phone = values[4]
        address = values[5]
        try:
            district = District.objects.get(name=district)
        except:
            district = District.objects.create(name=district)
        try:
            school = School.objects.get(INN=INN)
            school.district = district
            school.name = name
            school.shortname = shortname
            school.phone = phone
            school.address = address
            school.save()
        except BaseException as ex:
            logger.exception(ex)


class ExcelWriter(xlwt.Workbook):
    START_ROW = 4
    START_COLUMN = 1

    def __init__(self):
        super().__init__()
        self.row = START_ROW
        self.column = START_COLUMN
        self.shit: Worksheet = self.add_sheet("export", cell_overwrite_ok=True)
        self.schools = list(School.objects.all())
        self.buildings = list(Building.objects.all())
        self.zppp = list(ZPPP.objects.all())
        self.template_path = settings.DOCUMENT_ROOT + '/template.xlsx'
        self.full_path = settings.DOCUMENT_ROOT + "/export.xlsx"

    def add_buildings_to_excel(self, temperature_data: dict = None, data=None, full=False):
        global end, temperature_fields
        if full:
            fields = list(get_model_fields(Building))
        else:
            fields = self.filter_fields(data, Building)
        start = None
        if temperature_data is not None:
            if temperature_data.__contains__('start'):
                start = parse_date(temperature_data.pop('start'))
            if temperature_data.__contains__('end'):
                end = parse_date(temperature_data.pop('end'))
            else:
                end = datetime.date.today()
            if full:
                temperature_fields = list(get_model_fields(Temperature))
            else:
                temperature_fields = self.filter_fields(temperature_data, Temperature)
            try:
                temperature_data.pop('date')
            except:
                pass

        next_column = self.column
        self.row = self.START_ROW
        for school in self.schools:
            cur_column = self.column
            buildings = list(school.building_set.all())
            self.shit.write(1, cur_column, Building._meta.verbose_name)
            for building in buildings:
                for field in fields:
                    self.shit.write(2, cur_column, field.verbose_name.__str__())
                    self.shit.write(self.row, cur_column, getattr(building, field.name).__str__())
                    cur_column += 1
                if start is not None:
                    cur_column = self.add_building_temperature_to_excel(start=start, end=end, fields=temperature_fields,
                                                                        cur_column=cur_column,
                                                                        building=building)
            if len(buildings) == 0:
                self.shit.write(self.row, cur_column, "-")
            self.row += 1
            if next_column < cur_column:
                next_column = cur_column
        self.column = next_column

    def add_building_temperature_to_excel(self, start, end, fields, cur_column, building):
        while start <= end:
            self.shit.write(1, cur_column, start.__str__())
            try:
                obj = Temperature.objects.get(date=start, building=building)
            except:
                obj = None
            for field in fields:
                self.shit.write(2, cur_column, field.verbose_name.__str__())
                if obj is not None:
                    if getattr(obj, field.name) is not None:
                        self.shit.write(self.row, cur_column, getattr(obj, field.name).__str__())
                    else:
                        self.shit.write(self.row, cur_column, "-")
                else:
                    self.shit.write(self.row, cur_column, "-")
                cur_column += 1
            start += datetime.timedelta(days=1)

        return cur_column

    def add_foreign_model_to_excel(self, model: models.Model, to=None, data=None, full=False):
        if full:
            fields = list(get_model_fields(model))
        else:
            fields = self.filter_fields(data, model)
        next_column = self.column
        self.row = self.START_ROW
        get_atr_str = str(model._meta.model_name).lower()
        get_atr_str += "_set"
        for school in self.schools:
            cur_column = self.column
            if to is None:
                objects_list = list(school.__getattribute__(get_atr_str).all())
            self.shit.write(1, cur_column, model._meta.verbose_name)
            for model_object in objects_list:
                for field in fields:
                    self.shit.write(2, cur_column, field.verbose_name.__str__())
                    self.shit.write(self.row, cur_column, getattr(model_object, field.name).__str__())
                    cur_column += 1
            else:
                if len(objects_list) == 0:
                    self.shit.write(self.row, cur_column, "-")
            self.row += 1
            if next_column < cur_column:
                next_column = cur_column

        self.column = next_column

    def add_one_model_to_excel(self, model: models.Model, data=None, full=False):
        if full:
            fields = list(get_model_fields(model))
        else:
            fields = self.filter_fields(data, model)
        next_column = self.column
        self.row = self.START_ROW
        get_atr_str = "get_" + str(model._meta.model_name).lower()
        for school in self.schools:
            model_object = school.__getattribute__(get_atr_str)()
            cur_column = self.column
            for field in fields:
                self.shit.write(2, cur_column, field.verbose_name.__str__())
                self.shit.write(self.row, cur_column, getattr(model_object, field.name).__str__())
                cur_column += 1
            self.row += 1
            if next_column < cur_column:
                next_column = cur_column

        self.column = next_column

    def add_schools_to_excel(self, data=None, full=False):
        if full:
            fields = list(get_model_fields(School))
        else:
            fields = self.filter_fields(data, School)
        next_column = self.column
        self.row = self.START_ROW

        for model_object in self.schools:
            cur_column = self.column
            for field in fields:
                self.shit.write(2, cur_column, field.verbose_name.__str__())
                self.shit.write(self.row, cur_column, getattr(model_object, field.name).__str__())
                cur_column += 1
            self.row += 1
            if next_column < cur_column:
                next_column = cur_column

        self.column = next_column

    def filter_fields(self, data, model):
        if isinstance(data, str):
            data = json.loads(data)
        fields = list(get_model_fields(model))
        fields = list(filter(lambda f_field: data.__contains__(f_field.name), fields))
        fields = list(filter(lambda f_field: data[f_field.name], fields))
        return fields

    def filter_schools(self, filters: dict):
        if filters.__contains__(District._meta.model_name):
            cur_filter: dict = filters[District._meta.model_name]
            self.schools = filter(lambda f_school: cur_filter.__contains__(f_school.district.id.__str__()),
                                  self.schools)
            self.schools = filter(lambda f_school: cur_filter[f_school.district.id.__str__()], self.schools)

    def filter_buildings(self, filters: dict):
        if filters.__contains__(District._meta.model_name):
            cur_filter: dict = filters[District._meta.model_name]
            self.buildings = filter(lambda f_building: cur_filter.__contains__(f_building.school.district.id.__str__()),
                                    self.buildings)
            self.buildings = filter(lambda f_building: cur_filter[f_building.school.district.id.__str__()],
                                    self.buildings)

    def filter_zppp(self, filters: dict):
        cur_filter: dict = filters[District._meta.model_name]
        self.zppp = filter(lambda f_zppp: cur_filter.__contains__(f_zppp.school.district.id.__str__()),
                           self.zppp)
        self.zppp = filter(lambda f_zppp: cur_filter[f_zppp.school.district.id.__str__()], self.zppp)

    def filter_by_date_field(self, objects, start: datetime = None, end: datetime = None):
        if not start is None:
            objects = list(filter(lambda o: o.date >= start, objects))
        if not end is None:
            objects = list(filter(lambda o: o.date <= end, objects))
        return objects

    # def make_export_file(self, data: dict):
    #     if data.__contains__("filters"):
    #         self.filter_schools(data["filters"])
    #     self.schools = list(self.schools)
    #     if data.__contains__(School._meta.model_name):
    #         self.add_schools_to_excel(data=data[School._meta.model_name])
    #     if data.__contains__(Building._meta.model_name):
    #         if data.__contains__(Temperature._meta.model_name):
    #             self.add_buildings_to_excel(data=data[Building._meta.model_name],
    #                                         temperature_data=data[Temperature._meta.model_name])
    #         else:
    #             self.add_buildings_to_excel(data=data[Building._meta.model_name])
    #     if data.__contains__(Director._meta.model_name):
    #         self.add_one_model_to_excel(data=data[Director._meta.model_name], model=Director())
    #     self.save(self.full_path)
    #     return self.full_path

    # def make_full_export_file(self):
    #     self.add_schools_to_excel(full=True)
    #     self.add_foreign_model_to_excel(model=Building(), full=True)
    #     self.save(self.full_path)
    #     return self.full_path

    def make_full_export_file(self):
        shutil.copy(self.template_path, self.full_path)
        read_book = openpyxl.load_workbook(self.full_path)
        legal_worksheet = read_book.worksheets[0]
        for school in self.schools:
            values = []
            values.append(school.INN)
            values.append(school.district.name)
            values.append(school.form_type)
            values.append(school.edu_type)
            values.append(school.name)
            values.append(school.shortname)
            try:
                director = school.director
                values.append(f'{director.last_name} {director.first_name} {director.patronymic}')
                values.append(director.phone)
            except:
                values.append('-')
                values.append('-')
            try:
                zavhoz = school.zavhoz
                values.append(f'{zavhoz.last_name} {zavhoz.first_name} {zavhoz.patronymic}')
                values.append(zavhoz.phone)
            except:
                values.append('-')
                values.append('-')
            try:
                bookkeeper = school.bookkeeper
                values.append(f'{bookkeeper.last_name} {bookkeeper.first_name} {bookkeeper.patronymic}')
                values.append(bookkeeper.phone)
            except:
                values.append('-')
                values.append('-')
            try:
                requisites = school.requisites
                values.append(requisites.formation_date)
            except:
                values.append('-')
            buildings_outstanding = school.building_set.filter(type=Building.TYPE.FREE_STANDING)
            buildings_all = school.building_set.all()
            buildings_not_outstanding = school.building_set.filter(~Q(type=Building.TYPE.FREE_STANDING))
            values.append(len(buildings_outstanding))
            values.append(sum([i.building_square if i.building_square is not None else 0 for i in buildings_all]))
            values.append(
                sum([i.building_square if i.building_square is not None else 0 for i in buildings_not_outstanding]))
            legal_worksheet.append(values)
        building_worksheet = read_book.worksheets[1]
        for building in self.buildings:
            values2 = []
            values2.append(building.school.INN)
            values2.append(building.school.district.name)
            values2.append(building.school.form_type)
            values2.append(building.school.edu_type)
            values2.append(building.school.shortname)
            values2.append(f'{building.street} {building.street_number}')
            values2.append(building.type)
            values2.append(building.purpose)
            values2.append(building.construction_year)
            values2.append(building.building_square)
            values2.append(building.land_square)
            values2.append(building.number_of_storeys)
            values2.append(building.build_height)
            values2.append(building.occupancy_proj)
            values2.append(building.occupancy_fact)
            values2.append(building.arend_square)
            values2.append(building.arend_use_type)
            values2.append(building.unused_square)
            values2.append(building.repair_need_square)
            values2.append(building.technical_condition)
            values2.append(building.last_repair_year)
            building_worksheet.append(values2)
        go = read_book.worksheets[2]
        for z in self.zppp:
            values3 = []
            school = z.school
            values3.append(z.zppp_name_school)
            values3.append(school.INN)
            values3.append(z.type_ownership)
            values3.append(z.zppp_type)
            values3.append(z.zppp_adress)
            values3.append(z.zppp_group)
            values3.append(z.zppp_height)
            values3.append(z.zppp_square)
            values3.append(z.zppp_material)
            values3.append(z.zppp_heating)
            values3.append(z.zppp_water)
            values3.append(z.zppp_sewerage)
            values3.append(z.zppp_light)
            values3.append(z.zppp_ventilation)
            values3.append(z.zppp_additional)
            values3 = [i if i is not None else '-' for i in values3]
            values3 = [i if not isinstance(i, bool) else ('Есть' if i == True else 'Нет') for i in values3]
            go.append(values3)
        ventilation = read_book.worksheets[3]
        for building in self.buildings:
            values4 = []
            school = building.school
            values4.append(school.INN)
            values4.append(school.district.name)
            values4.append(school.shortname)
            values4.append(f'{building.street} {building.street_number}')
            values4.append(building.type)
            values4.append(building.purpose)
            if building.engineering_communication is None:
                values4.append(None)
                values4.append(None)
            else:
                values4.append(
                    building.engineering_communication.number_of_automatic_control_systems_for_the_air_handling_unit)
                values4.append(building.engineering_communication.technical_condition_of_the_ventilation_system)
            if building.indoor_areas is None:
                for i in range(28):
                    values4.append(None)
            else:
                values4.append(building.indoor_areas.food_block_exhaust_ventilation)
                values4.append(building.indoor_areas.food_block_exhaust_ventilation_is_workable)
                values4.append(building.indoor_areas.food_block_ventilation_type)
                values4.append(building.indoor_areas.food_block_supply_ventilation)
                values4.append(building.indoor_areas.food_block_air_heater_type)
                values4.append(building.indoor_areas.gym_exhaust_ventilation)
                values4.append(building.indoor_areas.gym_exhaust_ventilation_is_workable)
                values4.append(building.indoor_areas.gym_ventilation_type)
                values4.append(building.indoor_areas.gym_supply_ventilation)
                values4.append(building.indoor_areas.gym_air_heater_type)
                values4.append(building.indoor_areas.auditorium_exhaust_ventilation)
                values4.append(building.indoor_areas.auditorium_exhaust_ventilation_is_workable)
                values4.append(building.indoor_areas.auditorium_ventilation_type)
                values4.append(building.indoor_areas.auditorium_supply_ventilation)
                values4.append(building.indoor_areas.auditorium_air_heater_type)
                values4.append(building.indoor_areas.bathroom_exhaust_ventilation)
                values4.append(building.indoor_areas.bathroom_exhaust_ventilation_is_workable)
                values4.append(building.indoor_areas.bathroom_ventilation_type)
                values4.append(building.indoor_areas.laundry_exhaust_ventilation)
                values4.append(building.indoor_areas.laundry_exhaust_ventilation_is_workable)
                values4.append(building.indoor_areas.laundry_ventilation_type)
                values4.append(building.indoor_areas.laundry_supply_ventilation)
                values4.append(building.indoor_areas.laundry_air_heater_type)
                values4.append(building.indoor_areas.pool_exhaust_ventilation)
                values4.append(building.indoor_areas.pool_exhaust_ventilation_condition)
                values4.append(building.indoor_areas.pool_ventilation_type)
                values4.append(building.indoor_areas.pool_supply_ventilation)
                values4.append(building.indoor_areas.pool_air_heater_type)

            values4 = [i if i is not None else '-' for i in values4]
            values4 = [i if not isinstance(i, bool) else ('Да' if i == True else 'Нет') for i in values4]

            ventilation.append(values4)
        read_book.save(self.full_path)
        return self.full_path

    def make_export_file(self, data: dict):
        if data.__contains__("filters"):
            self.filter_schools(data["filters"])
            self.filter_buildings(data["filters"])
            self.filter_zppp(data["filters"])
        self.schools = list(self.schools)
        self.buildings = list(self.buildings)
        self.zppp = list(self.zppp)
        shutil.copy(self.template_path, self.full_path)
        read_book = openpyxl.load_workbook(self.full_path)
        legal_worksheet = read_book.worksheets[0]
        for school in self.schools:
            values = []
            values.append(school.INN)
            values.append(school.district.name)
            values.append(school.form_type)
            values.append(school.edu_type)
            values.append(school.name)
            values.append(school.shortname)
            try:
                director = school.director
                values.append(f'{director.last_name} {director.first_name} {director.patronymic}')
                values.append(director.phone)
            except:
                values.append('-')
                values.append('-')
            try:
                zavhoz = school.zavhoz
                values.append(f'{zavhoz.last_name} {zavhoz.first_name} {zavhoz.patronymic}')
                values.append(zavhoz.phone)
            except:
                values.append('-')
                values.append('-')
            try:
                bookkeeper = school.bookkeeper
                values.append(f'{bookkeeper.last_name} {bookkeeper.first_name} {bookkeeper.patronymic}')
                values.append(bookkeeper.phone)
            except:
                values.append('-')
                values.append('-')
            try:
                requisites = school.requisites
                values.append(requisites.formation_date)
            except:
                values.append('-')
            buildings_outstanding = school.building_set.filter(type=Building.TYPE.FREE_STANDING)
            buildings_all = school.building_set.all()
            buildings_not_outstanding = school.building_set.filter(~Q(type=Building.TYPE.FREE_STANDING))
            values.append(len(buildings_outstanding))
            values.append(sum([i.building_square if i.building_square is not None else 0 for i in buildings_all]))
            values.append(
                sum([i.building_square if i.building_square is not None else 0 for i in buildings_not_outstanding]))
            legal_worksheet.append(values)
        building_worksheet = read_book.worksheets[1]
        for building in self.buildings:
            values2 = []
            values2.append(building.school.INN)
            values2.append(building.school.district.name)
            values2.append(building.school.form_type)
            values2.append(building.school.edu_type)
            values2.append(building.school.shortname)
            values2.append(f'{building.street} {building.street_number}')
            values2.append(building.type)
            values2.append(building.purpose)
            values2.append(building.construction_year)
            values2.append(building.building_square)
            values2.append(building.land_square)
            values2.append(building.number_of_storeys)
            values2.append(building.build_height)
            values2.append(building.occupancy_proj)
            values2.append(building.occupancy_fact)
            values2.append(building.arend_square)
            values2.append(building.arend_use_type)
            values2.append(building.unused_square)
            values2.append(building.repair_need_square)
            values2.append(building.technical_condition)
            values2.append(building.last_repair_year)
            building_worksheet.append(values2)
        go = read_book.worksheets[2]
        for z in self.zppp:
            values3 = []
            school = z.school
            values3.append(school.INN)
            values3.append(z.zppp_name_school)
            values3.append(z.type_ownership)
            values3.append(z.zppp_type)
            values3.append(z.zppp_adress)
            values3.append(z.zppp_group)
            values3.append(z.zppp_height)
            values3.append(z.zppp_square)
            values3.append(z.zppp_material)
            values3.append(z.zppp_heating)
            values3.append(z.zppp_water)
            values3.append(z.zppp_sewerage)
            values3.append(z.zppp_light)
            values3.append(z.zppp_ventilation)
            values3.append(z.zppp_additional)
            values3 = [i if i is not None else '-' for i in values3]
            values3 = [i if not isinstance(i, bool) else ('Есть' if i == True else 'Нет') for i in values3]
            go.append(values3)
        ventilation = read_book.worksheets[3]
        for building in self.buildings:
            values4 = []
            school = building.school
            values4.append(school.INN)
            values4.append(school.district.name)
            values4.append(school.shortname)
            values4.append(f'{building.street} {building.street_number}')
            values4.append(building.type)
            values4.append(building.purpose)
            if building.engineering_communication is None:
                values4.append(None)
                values4.append(None)
            else:
                values4.append(building.engineering_communication.number_of_automatic_control_systems_for_the_air_handling_unit)
                values4.append(building.engineering_communication.technical_condition_of_the_ventilation_system)
            if building.indoor_areas is None:
                for i in range(28):
                    values4.append(None)
            else:
                values4.append(building.indoor_areas.food_block_exhaust_ventilation)
                values4.append(building.indoor_areas.food_block_exhaust_ventilation_is_workable)
                values4.append(building.indoor_areas.food_block_ventilation_type)
                values4.append(building.indoor_areas.food_block_supply_ventilation)
                values4.append(building.indoor_areas.food_block_air_heater_type)
                values4.append(building.indoor_areas.gym_exhaust_ventilation)
                values4.append(building.indoor_areas.gym_exhaust_ventilation_is_workable)
                values4.append(building.indoor_areas.gym_ventilation_type)
                values4.append(building.indoor_areas.gym_supply_ventilation)
                values4.append(building.indoor_areas.gym_air_heater_type)
                values4.append(building.indoor_areas.auditorium_exhaust_ventilation)
                values4.append(building.indoor_areas.auditorium_exhaust_ventilation_is_workable)
                values4.append(building.indoor_areas.auditorium_ventilation_type)
                values4.append(building.indoor_areas.auditorium_supply_ventilation)
                values4.append(building.indoor_areas.auditorium_air_heater_type)
                values4.append(building.indoor_areas.bathroom_exhaust_ventilation)
                values4.append(building.indoor_areas.bathroom_exhaust_ventilation_is_workable)
                values4.append(building.indoor_areas.bathroom_ventilation_type)
                values4.append(building.indoor_areas.laundry_exhaust_ventilation)
                values4.append(building.indoor_areas.laundry_exhaust_ventilation_is_workable)
                values4.append(building.indoor_areas.laundry_ventilation_type)
                values4.append(building.indoor_areas.laundry_supply_ventilation)
                values4.append(building.indoor_areas.laundry_air_heater_type)
                values4.append(building.indoor_areas.pool_exhaust_ventilation)
                values4.append(building.indoor_areas.pool_exhaust_ventilation_condition)
                values4.append(building.indoor_areas.pool_ventilation_type)
                values4.append(building.indoor_areas.pool_supply_ventilation)
                values4.append(building.indoor_areas.pool_air_heater_type)

            values4 = [i if i is not None else '-' for i in values4]
            values4 = [i if not isinstance(i, bool) else ('Да' if i == True else 'Нет') for i in values4]

            ventilation.append(values4)
        read_book.save(self.full_path)
        return self.full_path


def make_export_file(data: dict):
    excel = ExcelWriter()
    return excel.make_export_file(data)

    # full_path = settings.DOCUMENT_ROOT + "/export.xls"
    # excel = xlwt.Workbook()
    # shit: Worksheet = excel.add_sheet("export", cell_overwrite_ok=True)
    # schools = School.objects.all()
    # districts = District.objects.all()
    # row = START_ROW
    #
    # column = START_COLUMN
    # next_column = column
    #
    # if data.__contains__(School._meta.model_name):
    #     cur_data = data[School._meta.model_name]
    #     if isinstance(cur_data, str):
    #         cur_data = json.loads(cur_data)
    #     fields = list(get_model_fields(School))
    #     fields = filter(lambda filter_field: cur_data.__contains__(filter_field.name), fields)
    #     fields = list(filter(lambda filter_field: cur_data[filter_field.name], fields))
    #
    #     for school in schools:
    #         cur_column = column
    #
    #         for field in fields:
    #             shit.write(2, cur_column, field.verbose_name.__str__())
    #             shit.write(row, cur_column, getattr(school, field.name).__str__())
    #             cur_column += 1
    #         row += 1
    #         if next_column < cur_column:
    #             next_column = cur_column
    #
    # column = next_column
    # row = START_ROW
    #
    # if data.__contains__(Building._meta.model_name):
    #     cur_data = data[Building._meta.model_name]
    #     if isinstance(cur_data, str):
    #         cur_data = json.loads(cur_data)
    #     fields = list(get_model_fields(Building))
    #     fields = filter(lambda field: cur_data.__contains__(field.name), fields)
    #     fields = list(filter(lambda field: cur_data[field.name], fields))
    #
    #     for school in schools:
    #         cur_column = column
    #         for building in school.building_set.all():
    #             shit.write(1, cur_column, "Здание")
    #             for field in fields:
    #                 shit.write(2, cur_column, field.verbose_name.__str__())
    #                 shit.write(row, cur_column, getattr(building, field.name).__str__())
    #                 cur_column += 1
    #         if cur_column > next_column:
    #             next_column = cur_column
    #         row += 1
    #
    # excel.save(full_path)
    # return full_path


def make_full_export_file():
    excel = ExcelWriter()
    return excel.make_full_export_file()

    # full_path = settings.DOCUMENT_ROOT + "/export.xls"
    # excel = xlwt.Workbook()
    # shit: Worksheet = excel.add_sheet("export", cell_overwrite_ok=True)
    # schools = School.objects.all()
    # districts = District.objects.all()
    # row = START_ROW
    # fields = get_model_fields(School)
    # fields = fields[1:]
    # column = 1
    # cur_column = column
    # for school in schools:
    #     cur_column = column
    #     for field in fields:
    #         shit.write(2, cur_column, field.verbose_name.__str__())
    #         shit.write(row, cur_column, getattr(school, field.name).__str__())
    #         cur_column += 1
    #     row += 1
    #
    # column = cur_column + 2
    # row = START_ROW
    # max_count = 0
    #
    # fields = list(get_model_fields(Building))
    # for field in fields:
    #     if isinstance(field, models.ForeignKey):
    #         fields.remove(field)
    # for school in schools:
    #     if max_count < len(school.building_set.all()):
    #         max_count = len(school.building_set.all())
    # for school in schools:
    #     cur_column = column
    #     for building in school.building_set.all():
    #         shit.write(1, cur_column, building._meta.verbose_name)
    #         for field in fields:
    #             shit.write(2, cur_column, field.verbose_name.__str__())
    #             if getattr(building, field.name) is not None:
    #                 shit.write(row, cur_column, getattr(building, field.name).__str__())
    #             else:
    #                 shit.write(row, cur_column, "-")
    #             cur_column += 1
    #     row += 1
    #
    # excel.save(full_path)
    # return full_path
